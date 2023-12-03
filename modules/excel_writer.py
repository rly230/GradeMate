import pandas as pd
from openpyxl import load_workbook


def write_to_excel(grading_data, file_name, sheet_name):
    """
    採点結果をExcelファイルに書き込む。

    :param grading_data: 学籍番号と採点結果を含む辞書。
    :param file_name: Excelファイルの名前。
    :param sheet_name: シートの名前。
    """
    df = pd.DataFrame(list(grading_data.items()), columns=["Student ID", "Score"])

    try:
        # 既存のExcelファイルにアクセス
        book = load_workbook(file_name)
        with pd.ExcelWriter(file_name, engine="openpyxl", mode="a") as writer:
            # 既存のシートに追記
            if sheet_name in book.sheetnames:
                startrow = book[sheet_name].max_row
            else:
                startrow = 0
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
                startrow=startrow,
            )
            # writer.save() の呼び出しは不要
    except FileNotFoundError:
        # ファイルが存在しない場合は新規作成
        df.to_excel(file_name, sheet_name=sheet_name, index=False)
